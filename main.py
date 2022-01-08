import os
from typing import Optional
from io import BytesIO
from fastapi import FastAPI, File, UploadFile
from arcgis import GIS
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

from sdg_connectors.uscensus import USCensus as usc

app = FastAPI()


@app.get("/")
def read_root():
  token = os.environ['SAMPLE_TOKEN']
  gis = GIS(token=token)
  return gis.users.me.username


@app.get("/uscensus")
def read_uscensus():
  return usc.census_time()


@app.get("/items/{item_id}")
def read_item(item_id: int, q: Optional[str] = None):
  return {"item_id": item_id, "q": q}

@app.post("/uploadfile/")
def create_upload_file(file: UploadFile = File(...)):
  wb2 = load_workbook(BytesIO(file.file.read()))
  sheet_names = wb2.sheetnames


  # write to disk
  # with open(file.filename, "wb+") as f:
  #   f.write(file.file.read())

  return {"sheet_names": sheet_names}
